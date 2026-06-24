"""
App de Pedidos Coteminas - Flask local
Gestión de pedidos mensuales para el proveedor Coteminas.

Uso:
    python app.py
    Abrir http://localhost:5050 en el navegador
"""
import os
import io
import re
import json
from datetime import datetime
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file
)
from models import get_db, init_db
from parser_confirmacion import parse_email

app = Flask(__name__)
app.secret_key = 'pedidos-coteminas-tutextil-2026'

MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

CATEGORIAS_ORDEN = [
    'toalla', 'toallon', 'toallon_grande', 'juego', 'piso',
    'visita', 'toallita', 'bata', 'repasador',
    'sabana', 'funda', 'almohada', 'colcha',
    'acolchado', 'protector', 'mantel', 'camino', 'cortina', 'cubre',
]

CATEGORIAS_DISPLAY = {
    'toalla': 'Toallas',
    'toallon': 'Toallones',
    'toallon_grande': 'Toallones Grandes',
    'juego': 'Juegos',
    'piso': 'Toalla Piso',
    'visita': 'Toalla Visita',
    'toallita': 'Toallitas',
    'bata': 'Batas',
    'repasador': 'Repasadores',
    'sabana': 'Sábanas',
    'funda': 'Fundas',
    'almohada': 'Almohadas',
    'colcha': 'Colchas',
    'acolchado': 'Acolchados',
    'protector': 'Protectores',
    'mantel': 'Manteles',
    'camino': 'Caminos',
    'cortina': 'Cortinas',
    'cubre': 'Cubre',
}


# Keywords para detectar la categoría de un item manual a partir de su descripción.
# Orden importa: las más específicas primero (ej "TOALLON G" antes de "TOALLON").
CATEGORIA_KEYWORDS = [
    ('REPASADOR', 'repasador'),
    ('TOALLON G', 'toallon_grande'),
    ('TOALLON', 'toallon'),
    ('TOALLÓN', 'toallon'),
    ('TOALLA PISO', 'piso'),
    ('TOALLITA', 'toallita'),
    ('PISO', 'piso'),
    ('TOALLA', 'toalla'),
    ('JUEGO', 'juego'),
    ('JGO', 'juego'),
    ('BATA', 'bata'),
    ('VISITA', 'visita'),
    ('SABANA', 'sabana'),
    ('SÁBANA', 'sabana'),
    ('ALMOHADA', 'almohada'),
    ('FUNDA', 'funda'),
    ('COLCHA', 'colcha'),
    ('ACOLCHADO', 'acolchado'),
    ('PROTECTOR', 'protector'),
    ('MANTEL', 'mantel'),
    ('CAMINO', 'camino'),
    ('CORTINA', 'cortina'),
    ('CUBRE', 'cubre'),
]


def detectar_categoria_desc(descripcion):
    """Detectar la categoría de un item a partir de su descripción.
    Devuelve la categoría o None si no se reconoce."""
    if not descripcion:
        return None
    desc_upper = descripcion.upper()
    # Primero por inicio de la descripción (más confiable)
    for keyword, cat in CATEGORIA_KEYWORDS:
        if desc_upper.startswith(keyword):
            return cat
    # Luego por aparición en cualquier parte
    for keyword, cat in CATEGORIA_KEYWORDS:
        if keyword in desc_upper:
            return cat
    return None


@app.template_filter('mes_nombre')
def mes_nombre_filter(mes):
    return MESES.get(mes, str(mes))


@app.template_filter('cat_display')
def cat_display_filter(cat):
    return CATEGORIAS_DISPLAY.get(cat, cat.title() if cat else 'Otro')


def _etiqueta_tipo_pedido(pedido):
    """Devuelve 'Total' o 'Parcial N' para un pedido.
    Los parciales se numeran según el orden de creación dentro del mismo mes/año."""
    if pedido['tipo'] == 'total':
        return 'Total'
    # Parcial: numerar según cuántos parciales del mismo mes/año hay hasta este (inclusive)
    conn = get_db()
    parciales = conn.execute("""
        SELECT id FROM pedidos
        WHERE tipo = 'parcial' AND mes = ? AND anio = ?
        ORDER BY created_at, id
    """, (pedido['mes'], pedido['anio'])).fetchall()
    conn.close()
    ids = [r['id'] for r in parciales]
    try:
        num = ids.index(pedido['id']) + 1
    except ValueError:
        num = len(ids) + 1
    return f'Parcial {num}'


@app.template_filter('etiqueta_tipo')
def etiqueta_tipo_filter(pedido):
    return _etiqueta_tipo_pedido(pedido)


@app.template_filter('fecha_legible')
def fecha_legible_filter(fecha):
    """Convierte 'YYYY-MM-DD' (o timestamp) a 'DD/MM/YYYY'."""
    if not fecha:
        return ''
    try:
        f = str(fecha)[:10]  # tomar solo YYYY-MM-DD
        partes = f.split('-')
        if len(partes) == 3:
            return f"{partes[2]}/{partes[1]}/{partes[0]}"
    except Exception:
        pass
    return str(fecha)


# ── PÁGINA PRINCIPAL ─────────────────────────────────────────────────

@app.route('/')
def index():
    """Dashboard principal con pedidos recientes."""
    conn = get_db()
    pedidos = conn.execute("""
        SELECT p.*,
            (SELECT COUNT(*) FROM pedido_items WHERE pedido_id = p.id AND destino='tucuman') as items_tucuman,
            (SELECT COUNT(*) FROM pedido_items WHERE pedido_id = p.id AND destino='alsina') as items_alsina,
            (SELECT SUM(cantidad) FROM pedido_items WHERE pedido_id = p.id AND destino='tucuman') as total_tucuman,
            (SELECT SUM(cantidad) FROM pedido_items WHERE pedido_id = p.id AND destino='alsina') as total_alsina
        FROM pedidos p
        ORDER BY p.anio DESC, p.mes DESC, p.created_at DESC
    """).fetchall()
    conn.close()
    return render_template('index.html', pedidos=pedidos, meses=MESES)


# ── PEDIDOS ──────────────────────────────────────────────────────────

@app.route('/pedido/nuevo', methods=['GET', 'POST'])
def nuevo_pedido():
    """Crear un nuevo pedido."""
    if request.method == 'POST':
        mes = int(request.form['mes'])
        anio = int(request.form['anio'])
        tipo = request.form.get('tipo', 'total')
        fecha_pedido = request.form.get('fecha_pedido', '').strip() or datetime.now().strftime('%Y-%m-%d')
        notas = request.form.get('notas', '').strip()

        conn = get_db()
        cursor = conn.execute(
            "INSERT INTO pedidos (mes, anio, tipo, estado, fecha_pedido, notas) VALUES (?, ?, ?, 'sin_finalizar', ?, ?)",
            (mes, anio, tipo, fecha_pedido, notas)
        )
        pedido_id = cursor.lastrowid
        conn.commit()
        conn.close()

        flash(f'Pedido {MESES[mes]} {anio} creado', 'success')
        return redirect(url_for('editar_pedido', pedido_id=pedido_id))

    now = datetime.now()
    return render_template('nuevo_pedido.html',
                           meses=MESES,
                           anio_actual=now.year,
                           mes_actual=now.month,
                           fecha_hoy=now.strftime('%Y-%m-%d'))


@app.route('/pedido/<int:pedido_id>')
def editar_pedido(pedido_id):
    """Pantalla principal de edición de un pedido."""
    conn = get_db()
    pedido = conn.execute("SELECT * FROM pedidos WHERE id = ?", (pedido_id,)).fetchone()
    if not pedido:
        flash('Pedido no encontrado', 'error')
        return redirect(url_for('index'))

    # Items del pedido agrupados por destino
    items_tucuman = conn.execute("""
        SELECT pi.*, p.descripcion as prod_descripcion, COALESCE(p.categoria, pi.categoria_manual) as prod_categoria,
               p.linea as prod_linea, p.ean as prod_ean,
               p.sku_tu_textil as prod_sku, p.mla_ids as prod_mla
        FROM pedido_items pi
        LEFT JOIN productos p ON pi.producto_id = p.id
        WHERE pi.pedido_id = ? AND pi.destino = 'tucuman'
        ORDER BY p.categoria, p.linea, p.descripcion
    """, (pedido_id,)).fetchall()

    items_alsina = conn.execute("""
        SELECT pi.*, p.descripcion as prod_descripcion, COALESCE(p.categoria, pi.categoria_manual) as prod_categoria,
               p.linea as prod_linea, p.ean as prod_ean,
               p.sku_tu_textil as prod_sku, p.mla_ids as prod_mla
        FROM pedido_items pi
        LEFT JOIN productos p ON pi.producto_id = p.id
        WHERE pi.pedido_id = ? AND pi.destino = 'alsina'
        ORDER BY p.categoria, p.linea, p.descripcion
    """, (pedido_id,)).fetchall()

    # Categorías disponibles en el catálogo
    categorias = conn.execute("""
        SELECT DISTINCT categoria FROM productos WHERE estado='activo' ORDER BY categoria
    """).fetchall()

    conn.close()

    return render_template('editar_pedido.html',
                           pedido=pedido,
                           items_tucuman=items_tucuman,
                           items_alsina=items_alsina,
                           categorias=categorias,
                           categorias_display=CATEGORIAS_DISPLAY,
                           meses=MESES)


@app.route('/pedido/<int:pedido_id>/estado', methods=['POST'])
def cambiar_estado_pedido(pedido_id):
    """Cambiar estado del pedido (sin_finalizar <-> finalizado)."""
    conn = get_db()
    pedido = conn.execute("SELECT * FROM pedidos WHERE id = ?", (pedido_id,)).fetchone()
    if not pedido:
        flash('Pedido no encontrado', 'error')
        return redirect(url_for('index'))

    nuevo_estado = request.form.get('estado', 'sin_finalizar')
    cerrado_at = datetime.now().isoformat() if nuevo_estado == 'finalizado' else None

    conn.execute("""
        UPDATE pedidos SET estado = ?, cerrado_at = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (nuevo_estado, cerrado_at, pedido_id))
    conn.commit()
    conn.close()

    if nuevo_estado == 'finalizado':
        flash('Pedido marcado como FINALIZADO', 'success')
    else:
        flash('Pedido marcado como SIN FINALIZAR (podés seguir agregando)', 'info')

    return redirect(url_for('editar_pedido', pedido_id=pedido_id))


@app.route('/pedido/<int:pedido_id>/editar-info', methods=['POST'])
def editar_info_pedido(pedido_id):
    """Editar mes, año y tipo de un pedido existente."""
    conn = get_db()
    pedido = conn.execute("SELECT * FROM pedidos WHERE id = ?", (pedido_id,)).fetchone()
    if not pedido:
        flash('Pedido no encontrado', 'error')
        return redirect(url_for('index'))

    mes = int(request.form.get('mes', pedido['mes']))
    anio = int(request.form.get('anio', pedido['anio']))
    tipo = request.form.get('tipo', pedido['tipo'])
    fecha_pedido = request.form.get('fecha_pedido', '').strip() or pedido['fecha_pedido']
    notas = request.form.get('notas', pedido['notas'] or '').strip()

    conn.execute("""
        UPDATE pedidos SET mes = ?, anio = ?, tipo = ?, fecha_pedido = ?, notas = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (mes, anio, tipo, fecha_pedido, notas, pedido_id))
    conn.commit()
    conn.close()

    flash('Datos del pedido actualizados', 'success')
    return redirect(url_for('editar_pedido', pedido_id=pedido_id))


@app.route('/pedido/<int:pedido_id>/eliminar', methods=['POST'])
def eliminar_pedido(pedido_id):
    """Eliminar un pedido y todos sus items."""
    conn = get_db()
    conn.execute("DELETE FROM pedido_items WHERE pedido_id = ?", (pedido_id,))
    conn.execute("DELETE FROM confirmacion_items WHERE confirmacion_id IN (SELECT id FROM confirmaciones WHERE pedido_id = ?)", (pedido_id,))
    conn.execute("DELETE FROM confirmaciones WHERE pedido_id = ?", (pedido_id,))
    conn.execute("DELETE FROM pedidos WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()
    flash('Pedido eliminado', 'info')
    return redirect(url_for('index'))


# ── CATÁLOGO / BÚSQUEDA DE PRODUCTOS ─────────────────────────────────

@app.route('/api/productos')
def api_productos():
    """API para buscar productos del catálogo. Usado por AJAX."""
    q = request.args.get('q', '').strip()
    categoria = request.args.get('categoria', '').strip()
    linea = request.args.get('linea', '').strip()
    limit = int(request.args.get('limit', 50))

    conn = get_db()
    conditions = ["estado = 'activo'"]
    params = []

    if q:
        # Buscar en descripción, código, EAN, SKU interno y MLA
        conditions.append("(descripcion LIKE ? OR codigo LIKE ? OR ean LIKE ? OR sku_tu_textil LIKE ? OR mla_ids LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like, like, like, like])

    if categoria:
        conditions.append("categoria = ?")
        params.append(categoria)

    if linea:
        conditions.append("linea = ?")
        params.append(linea)

    where = " AND ".join(conditions)
    rows = conn.execute(f"""
        SELECT id, codigo, descripcion, categoria, linea, diseno,
               color_nombre, color_codigo, ean, tamanio, fuente,
               sku_tu_textil, mla_ids
        FROM productos
        WHERE {where}
        ORDER BY categoria, linea, descripcion
        LIMIT ?
    """, params + [limit]).fetchall()

    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/lineas')
def api_lineas():
    """Obtener líneas disponibles, opcionalmente filtradas por categoría."""
    categoria = request.args.get('categoria', '').strip()
    conn = get_db()
    if categoria:
        rows = conn.execute("""
            SELECT DISTINCT linea FROM productos
            WHERE estado='activo' AND categoria=? AND linea IS NOT NULL
            ORDER BY linea
        """, (categoria,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT DISTINCT linea FROM productos
            WHERE estado='activo' AND linea IS NOT NULL
            ORDER BY linea
        """).fetchall()
    conn.close()
    return jsonify([r['linea'] for r in rows])


@app.route('/api/familias')
def api_familias():
    """Obtener familias (agrupaciones por categoría+línea+diseño) para selección agrupada.
    Devuelve familias con la cantidad de colores/variantes disponibles."""
    q = request.args.get('q', '').strip()
    categoria = request.args.get('categoria', '').strip()
    linea = request.args.get('linea', '').strip()

    conn = get_db()
    conditions = ["estado = 'activo'"]
    params = []

    if q:
        conditions.append("(descripcion LIKE ? OR codigo LIKE ? OR ean LIKE ? OR sku_tu_textil LIKE ? OR mla_ids LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like, like, like, like])
    if categoria:
        conditions.append("categoria = ?")
        params.append(categoria)
    if linea:
        conditions.append("linea = ?")
        params.append(linea)

    where = " AND ".join(conditions)

    # Agrupar por SKU interno + categoría + línea + diseño (sin color)
    # Cada familia abarca todos los colores de un mismo modelo.
    rows = conn.execute(f"""
        SELECT categoria, linea, diseno, sku_tu_textil, mla_ids,
               COUNT(*) as variantes,
               GROUP_CONCAT(DISTINCT color_nombre) as colores,
               MIN(id) as primer_id
        FROM productos
        WHERE {where} AND diseno IS NOT NULL
        GROUP BY categoria, linea, diseno, sku_tu_textil
        HAVING COUNT(*) > 1
        ORDER BY categoria, linea, diseno
        LIMIT 40
    """, params).fetchall()

    conn.close()

    familias = []
    for r in rows:
        colores_list = [c for c in (r['colores'].split(',') if r['colores'] else []) if c]
        cat_display = CATEGORIAS_DISPLAY.get(r['categoria'], r['categoria'] or '')
        nombre = f"{cat_display} {r['linea'] or ''} {r['diseno'] or ''}".strip()
        familias.append({
            'nombre': nombre,
            'categoria': r['categoria'],
            'linea': r['linea'],
            'diseno': r['diseno'],
            'sku_tu_textil': r['sku_tu_textil'],
            'mla_ids': r['mla_ids'],
            'variantes': r['variantes'],
            'colores': colores_list,
            'colores_str': ', '.join(colores_list),
        })

    return jsonify(familias)


# ── ITEMS DEL PEDIDO ─────────────────────────────────────────────────

@app.route('/pedido/<int:pedido_id>/agregar', methods=['POST'])
def agregar_item(pedido_id):
    """Agregar item(s) al pedido. Soporta múltiples items vía JSON."""
    conn = get_db()
    pedido = conn.execute("SELECT * FROM pedidos WHERE id = ?", (pedido_id,)).fetchone()
    if not pedido:
        return jsonify({'error': 'Pedido no encontrado'}), 404

    # Detectar si viene como JSON (multi-item) o form (legacy single)
    if request.is_json:
        items = request.get_json()
        if not isinstance(items, list):
            items = [items]

        count = 0
        for item_data in items:
            es_manual = item_data.get('es_manual', False)
            destino = item_data.get('destino', 'tucuman')
            cantidad = int(item_data.get('cantidad', 0))
            color_detalle = item_data.get('color_detalle', '').strip()
            notas = item_data.get('notas', '').strip()

            if es_manual:
                desc = item_data.get('descripcion_manual', '').strip()
                codigo = item_data.get('codigo_manual', '').strip()
                cat_manual = (item_data.get('categoria_manual') or '').strip() or None
                # Si no vino categoría, intentar detectarla de la descripción
                if not cat_manual:
                    cat_manual = detectar_categoria_desc(desc)
                es_familia = 1 if item_data.get('es_familia') else 0
                if desc:
                    conn.execute("""
                        INSERT INTO pedido_items (pedido_id, destino, cantidad, color_detalle, notas,
                                                  es_manual, descripcion_manual, codigo_manual,
                                                  categoria_manual, es_familia)
                        VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?)
                    """, (pedido_id, destino, cantidad, color_detalle, notas, desc, codigo,
                          cat_manual, es_familia))
                    count += 1
            else:
                producto_id = item_data.get('producto_id')
                if producto_id:
                    conn.execute("""
                        INSERT INTO pedido_items (pedido_id, producto_id, destino, cantidad,
                                                  color_detalle, notas)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (pedido_id, int(producto_id), destino, cantidad, color_detalle, notas))
                    count += 1

        conn.execute("UPDATE pedidos SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (pedido_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'count': count})

    # Fallback: form POST (single item, legacy)
    es_manual = request.form.get('es_manual') == '1'
    destino = request.form.get('destino', 'tucuman')
    cantidad = int(request.form.get('cantidad', 0))
    color_detalle = request.form.get('color_detalle', '').strip()
    notas = request.form.get('notas', '').strip()

    if es_manual:
        descripcion_manual = request.form.get('descripcion_manual', '').strip()
        codigo_manual = request.form.get('codigo_manual', '').strip()
        categoria_manual = (request.form.get('categoria_manual') or '').strip() or None
        # Si no eligió categoría, detectarla de la descripción
        if not categoria_manual:
            categoria_manual = detectar_categoria_desc(descripcion_manual)
        if not descripcion_manual:
            return jsonify({'error': 'Descripción requerida para item manual'}), 400

        conn.execute("""
            INSERT INTO pedido_items (pedido_id, destino, cantidad, color_detalle, notas,
                                      es_manual, descripcion_manual, codigo_manual, categoria_manual)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
        """, (pedido_id, destino, cantidad, color_detalle, notas,
              descripcion_manual, codigo_manual, categoria_manual))
    else:
        producto_id = request.form.get('producto_id')
        if not producto_id:
            return jsonify({'error': 'Producto requerido'}), 400

        conn.execute("""
            INSERT INTO pedido_items (pedido_id, producto_id, destino, cantidad,
                                      color_detalle, notas)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido_id, int(producto_id), destino, cantidad, color_detalle, notas))

    conn.execute("UPDATE pedidos SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()

    flash('Item agregado', 'success')
    return redirect(url_for('editar_pedido', pedido_id=pedido_id))


@app.route('/pedido/<int:pedido_id>/item/<int:item_id>/editar', methods=['POST'])
def editar_item(pedido_id, item_id):
    """Editar un item del pedido."""
    conn = get_db()
    cantidad = int(request.form.get('cantidad', 0))
    color_detalle = request.form.get('color_detalle', '').strip()
    notas = request.form.get('notas', '').strip()
    destino = request.form.get('destino')
    # categoria_manual: solo se aplica a items manuales (los del catálogo usan la del producto)
    categoria_manual = request.form.get('categoria_manual')

    updates = ["cantidad = ?", "color_detalle = ?", "notas = ?"]
    params = [cantidad, color_detalle, notas]

    if destino:
        updates.append("destino = ?")
        params.append(destino)

    if categoria_manual is not None:
        cat_val = categoria_manual.strip() or None
        # Si quedó vacío (Detectar automáticamente), deducir de la descripción del item
        if not cat_val:
            row = conn.execute(
                "SELECT descripcion_manual FROM pedido_items WHERE id = ? AND pedido_id = ?",
                (item_id, pedido_id)
            ).fetchone()
            if row and row['descripcion_manual']:
                cat_val = detectar_categoria_desc(row['descripcion_manual'])
        updates.append("categoria_manual = ?")
        params.append(cat_val)

    params.extend([item_id, pedido_id])
    conn.execute(f"""
        UPDATE pedido_items SET {', '.join(updates)}
        WHERE id = ? AND pedido_id = ?
    """, params)

    conn.execute("UPDATE pedidos SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()
    flash('Item actualizado', 'success')
    return redirect(url_for('editar_pedido', pedido_id=pedido_id))


@app.route('/pedido/<int:pedido_id>/item/<int:item_id>/eliminar', methods=['POST'])
def eliminar_item(pedido_id, item_id):
    """Eliminar un item del pedido."""
    conn = get_db()
    conn.execute("DELETE FROM pedido_items WHERE id = ? AND pedido_id = ?", (item_id, pedido_id))
    conn.execute("UPDATE pedidos SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (pedido_id,))
    conn.commit()
    conn.close()
    flash('Item eliminado', 'info')
    return redirect(url_for('editar_pedido', pedido_id=pedido_id))


# ── RESUMEN DEL PEDIDO ───────────────────────────────────────────────

@app.route('/pedido/<int:pedido_id>/resumen')
def resumen_pedido(pedido_id):
    """Resumen detallado del pedido por depósito."""
    conn = get_db()
    pedido = conn.execute("SELECT * FROM pedidos WHERE id = ?", (pedido_id,)).fetchone()
    if not pedido:
        flash('Pedido no encontrado', 'error')
        return redirect(url_for('index'))

    resumen = {}
    for destino in ['tucuman', 'alsina']:
        items = conn.execute("""
            SELECT pi.*, p.descripcion as prod_descripcion, COALESCE(p.categoria, pi.categoria_manual) as prod_categoria,
                   p.linea as prod_linea, p.ean as prod_ean, p.diseno as prod_diseno,
                   p.color_nombre as prod_color
            FROM pedido_items pi
            LEFT JOIN productos p ON pi.producto_id = p.id
            WHERE pi.pedido_id = ? AND pi.destino = ?
            ORDER BY p.categoria, p.linea, p.descripcion
        """, (pedido_id, destino)).fetchall()

        # Agrupar por categoría
        por_categoria = {}
        total_unidades = 0
        for item in items:
            cat = item['prod_categoria'] if item['prod_categoria'] else 'manual'
            if cat not in por_categoria:
                por_categoria[cat] = {'items': [], 'subtotal': 0}
            por_categoria[cat]['items'].append(item)
            por_categoria[cat]['subtotal'] += item['cantidad'] or 0
            total_unidades += item['cantidad'] or 0

        # Ordenar categorías
        categorias_ordenadas = {}
        for cat in CATEGORIAS_ORDEN + ['manual']:
            if cat in por_categoria:
                categorias_ordenadas[cat] = por_categoria[cat]
        # Agregar las que no están en el orden
        for cat in por_categoria:
            if cat not in categorias_ordenadas:
                categorias_ordenadas[cat] = por_categoria[cat]

        resumen[destino] = {
            'items': items,
            'por_categoria': categorias_ordenadas,
            'total_unidades': total_unidades,
            'total_items': len(items),
        }

    # Confirmaciones
    confirmaciones = conn.execute("""
        SELECT * FROM confirmaciones WHERE pedido_id = ? ORDER BY fecha_confirmacion DESC
    """, (pedido_id,)).fetchall()

    conn.close()

    return render_template('resumen.html',
                           pedido=pedido,
                           resumen=resumen,
                           confirmaciones=confirmaciones,
                           meses=MESES,
                           categorias_display=CATEGORIAS_DISPLAY)


# ── EXPORTAR EXCEL ───────────────────────────────────────────────────

@app.route('/pedido/<int:pedido_id>/exportar')
def exportar_excel(pedido_id):
    """Exportar pedido a Excel con formato apilado (Alsina arriba, Tucumán abajo)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    conn = get_db()
    pedido = conn.execute("SELECT * FROM pedidos WHERE id = ?", (pedido_id,)).fetchone()
    if not pedido:
        flash('Pedido no encontrado', 'error')
        return redirect(url_for('index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pedido {MESES[pedido['mes']]} {pedido['anio']}"[:31]

    # Estilos
    cat_font = Font(bold=True, size=11)
    cat_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')

    # Anchos de columna: MLA | SKU | Descripción | Colores | Cantidad | Detalle
    ws.column_dimensions['A'].width = 20  # MLA
    ws.column_dimensions['B'].width = 14  # SKU
    ws.column_dimensions['C'].width = 48  # Descripción
    ws.column_dimensions['D'].width = 18  # Colores
    ws.column_dimensions['E'].width = 12  # Cantidad
    ws.column_dimensions['F'].width = 50  # Detalle

    sub_headers = ['MLA', 'SKU', 'DESCRIPCIÓN', 'COLORES', 'CANTIDAD', 'DETALLE']

    # Lookup SKU -> MLA (para items manuales de familias que guardan el SKU como código)
    sku_mla = {}
    for r in conn.execute("""
        SELECT DISTINCT sku_tu_textil, mla_ids FROM productos
        WHERE sku_tu_textil IS NOT NULL AND mla_ids IS NOT NULL AND mla_ids != ''
    """).fetchall():
        sku_mla[r['sku_tu_textil']] = r['mla_ids']

    # Obtener items por categoría para cada destino
    def get_items_por_cat(destino):
        items = conn.execute("""
            SELECT pi.*, p.descripcion as prod_descripcion, COALESCE(p.categoria, pi.categoria_manual) as prod_categoria,
                   p.linea as prod_linea, p.ean as prod_ean, p.codigo as prod_codigo,
                   p.color_nombre as prod_color, p.sku_tu_textil as prod_sku,
                   p.mla_ids as prod_mla
            FROM pedido_items pi
            LEFT JOIN productos p ON pi.producto_id = p.id
            WHERE pi.pedido_id = ? AND pi.destino = ?
            ORDER BY p.categoria, p.linea, p.descripcion
        """, (pedido_id, destino)).fetchall()

        por_cat = {}
        for item in items:
            cat = item['prod_categoria'] if item['prod_categoria'] else 'manual'
            por_cat.setdefault(cat, []).append(item)
        return por_cat

    def escribir_seccion(row_num, titulo, color_hex, cats_dict):
        """Escribe una sección completa (un depósito) en forma vertical.
        Devuelve (siguiente_fila, total_unidades)."""
        # Título del depósito
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        c = ws.cell(row=row_num, column=1, value=titulo)
        c.font = Font(bold=True, size=14, color='FFFFFF')
        c.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        c.alignment = Alignment(horizontal='center')
        row_num += 1

        # Sub-headers
        for i, h in enumerate(sub_headers):
            cell = ws.cell(row=row_num, column=i + 1, value=h)
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
        row_num += 1

        # Categorías presentes en este depósito
        cats = [c for c in (CATEGORIAS_ORDEN + ['manual']) if c in cats_dict]
        total = 0
        for cat in cats:
            cat_name = CATEGORIAS_DISPLAY.get(cat, cat.title()).upper()
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            cc = ws.cell(row=row_num, column=1, value=cat_name)
            cc.font = cat_font
            cc.fill = cat_fill
            row_num += 1

            for item in cats_dict[cat]:
                if item['es_manual']:
                    sku_m = item['codigo_manual'] or ''
                    ws.cell(row=row_num, column=1, value=sku_mla.get(sku_m, ''))
                    ws.cell(row=row_num, column=2, value=sku_m)
                    ws.cell(row=row_num, column=3, value=item['descripcion_manual'] or '')
                else:
                    ws.cell(row=row_num, column=1, value=item['prod_mla'] or '')
                    ws.cell(row=row_num, column=2, value=item['prod_sku'] or item['prod_codigo'] or '')
                    ws.cell(row=row_num, column=3, value=item['prod_descripcion'] or '')
                ws.cell(row=row_num, column=4, value=item['color_detalle'] or '')
                ws.cell(row=row_num, column=5, value=item['cantidad'] or 0)
                ws.cell(row=row_num, column=6, value=item['notas'] or '')
                # Bordes + wrap en detalle/colores
                for col in range(1, 7):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = thin_border
                    if col in (4, 6):
                        cell.alignment = wrap_align
                total += item['cantidad'] or 0
                row_num += 1

        # Total del depósito
        ws.cell(row=row_num, column=4, value='TOTAL:').font = Font(bold=True, size=12)
        tcell = ws.cell(row=row_num, column=5, value=total)
        tcell.font = Font(bold=True, size=12)
        row_num += 1
        return row_num, total

    alsina_cats = get_items_por_cat('alsina')
    tucuman_cats = get_items_por_cat('tucuman')

    # ── Título general con mes/tipo/fecha ──
    tipo_label = _etiqueta_tipo_pedido(pedido)
    ws.merge_cells('A1:F1')
    titulo = f"PEDIDO COTEMINAS — {MESES[pedido['mes']]} {pedido['anio']} — {tipo_label}"
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    fecha_txt = fecha_legible_filter(pedido['fecha_pedido'])
    ws['A2'] = f"Fecha del pedido: {fecha_txt}" if fecha_txt else ""
    ws['A2'].font = Font(size=11, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # ── ALSINA arriba, TUCUMÁN abajo ──
    row = 4
    row, total_alsina = escribir_seccion(row, 'ALSINA - ALBERTO', '2F5496', alsina_cats)
    row += 2  # Espacio entre depósitos
    row, total_tucuman = escribir_seccion(row, 'TUCUMÁN - SERGIO', '548235', tucuman_cats)

    conn.close()

    # Guardar a buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Nombre del archivo: PEDIDO COTEMINAS - Septiembre 2026 - Total / Parcial N
    filename = f"PEDIDO COTEMINAS - {MESES[pedido['mes']]} {pedido['anio']} - {tipo_label}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ── CONFIRMACIONES ───────────────────────────────────────────────────

@app.route('/pedido/<int:pedido_id>/confirmacion', methods=['POST'])
def agregar_confirmacion(pedido_id):
    """Pegar texto del email de confirmación del proveedor."""
    conn = get_db()
    texto = request.form.get('texto_email', '').strip()
    notas = request.form.get('notas', '').strip()

    if not texto:
        flash('Pegá el texto del email de confirmación', 'warning')
        return redirect(url_for('resumen_pedido', pedido_id=pedido_id))

    cursor = conn.execute("""
        INSERT INTO confirmaciones (pedido_id, texto_email, notas)
        VALUES (?, ?, ?)
    """, (pedido_id, texto, notas))
    conf_id = cursor.lastrowid

    # Intentar parsear el email: buscar productos mencionados
    # Por ahora guardar el texto crudo, el parseo se puede mejorar
    conn.commit()
    conn.close()

    flash('Confirmación guardada', 'success')
    return redirect(url_for('resumen_pedido', pedido_id=pedido_id))


# ── HISTORIAL ────────────────────────────────────────────────────────

@app.route('/historial')
def historial():
    """Ver historial de todos los pedidos."""
    conn = get_db()
    pedidos = conn.execute("""
        SELECT p.*,
            (SELECT SUM(cantidad) FROM pedido_items WHERE pedido_id = p.id AND destino='tucuman') as total_tucuman,
            (SELECT SUM(cantidad) FROM pedido_items WHERE pedido_id = p.id AND destino='alsina') as total_alsina,
            (SELECT SUM(cantidad) FROM pedido_items WHERE pedido_id = p.id) as total_general
        FROM pedidos p
        ORDER BY p.anio DESC, p.mes DESC, p.created_at DESC
    """).fetchall()
    conn.close()
    return render_template('historial.html', pedidos=pedidos, meses=MESES)


# ── CATÁLOGO COMPLETO ────────────────────────────────────────────────

@app.route('/catalogo')
def catalogo():
    """Ver el catálogo completo de productos."""
    conn = get_db()
    categoria = request.args.get('categoria', '')
    linea = request.args.get('linea', '')
    q = request.args.get('q', '')
    incluir_inactivos = request.args.get('inactivos') == '1'

    conditions = []
    params = []
    if not incluir_inactivos:
        conditions.append("estado = 'activo'")

    if q:
        conditions.append(
            "(descripcion LIKE ? OR codigo LIKE ? OR ean LIKE ? "
            "OR sku_tu_textil LIKE ? OR mla_ids LIKE ?)"
        )
        like = f"%{q}%"
        params.extend([like, like, like, like, like])
    if categoria:
        conditions.append("categoria = ?")
        params.append(categoria)
    if linea:
        conditions.append("linea = ?")
        params.append(linea)

    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    productos = conn.execute(f"""
        SELECT * FROM productos {where}
        ORDER BY categoria, linea, descripcion
        LIMIT 500
    """, params).fetchall()

    categorias = conn.execute("""
        SELECT DISTINCT categoria FROM productos WHERE estado='activo' ORDER BY categoria
    """).fetchall()

    lineas = conn.execute("""
        SELECT DISTINCT linea FROM productos WHERE estado='activo' AND linea IS NOT NULL ORDER BY linea
    """).fetchall()

    conn.close()
    return render_template('catalogo.html',
                           productos=productos,
                           categorias=categorias,
                           lineas=lineas,
                           filtro_categoria=categoria,
                           filtro_linea=linea,
                           filtro_q=q,
                           incluir_inactivos=incluir_inactivos,
                           categorias_display=CATEGORIAS_DISPLAY)


# Campos del catálogo que se pueden editar a mano desde la web
CAMPOS_PRODUCTO = [
    'codigo', 'descripcion', 'sku_tu_textil', 'mla_ids', 'categoria',
    'linea', 'diseno', 'color_nombre', 'color_codigo', 'tipo',
    'tamanio', 'ean', 'origen', 'coleccion',
]


def _leer_form_producto():
    """Lee los campos del producto desde request.form, normalizando vacíos a None."""
    datos = {}
    for campo in CAMPOS_PRODUCTO:
        val = (request.form.get(campo) or '').strip()
        datos[campo] = val or None
    # Normalizar MLA: quitar espacios y dejar separados por coma
    if datos.get('mla_ids'):
        partes = re.split(r'[\s,]+', datos['mla_ids'])
        datos['mla_ids'] = ','.join(p for p in partes if p)
    return datos


@app.route('/catalogo/nuevo', methods=['GET', 'POST'])
def producto_nuevo():
    """Crear un producto manualmente (por si falta algo del catálogo)."""
    if request.method == 'POST':
        datos = _leer_form_producto()
        if not datos.get('descripcion') or not datos.get('codigo'):
            flash('El código y la descripción son obligatorios.', 'warning')
            return render_template('producto_form.html', producto=datos, es_nuevo=True,
                                   categorias_display=CATEGORIAS_DISPLAY,
                                   categorias_orden=CATEGORIAS_ORDEN)
        conn = get_db()
        columnas = CAMPOS_PRODUCTO + ['estado', 'fuente']
        valores = [datos[c] for c in CAMPOS_PRODUCTO] + ['activo', 'manual']
        placeholders = ','.join('?' for _ in columnas)
        conn.execute(
            f"INSERT INTO productos ({','.join(columnas)}) VALUES ({placeholders})",
            valores
        )
        conn.commit()
        conn.close()
        flash(f"Producto «{datos['descripcion']}» creado.", 'success')
        return redirect(url_for('catalogo', q=datos.get('sku_tu_textil') or datos['codigo']))

    return render_template('producto_form.html', producto={}, es_nuevo=True,
                           categorias_display=CATEGORIAS_DISPLAY,
                           categorias_orden=CATEGORIAS_ORDEN)


@app.route('/catalogo/<int:producto_id>/duplicar', methods=['GET'])
def producto_duplicar(producto_id):
    """Abre el formulario de alta precargado con los datos de un producto existente.
    Al guardar crea un producto NUEVO (postea a /catalogo/nuevo), no modifica el original."""
    conn = get_db()
    producto = conn.execute("SELECT * FROM productos WHERE id = ?", (producto_id,)).fetchone()
    conn.close()
    if not producto:
        flash('No se encontró ese producto.', 'warning')
        return redirect(url_for('catalogo'))
    datos = {c: producto[c] for c in CAMPOS_PRODUCTO}
    return render_template('producto_form.html', producto=datos, es_nuevo=True,
                           duplicando=True, form_action=url_for('producto_nuevo'),
                           categorias_display=CATEGORIAS_DISPLAY,
                           categorias_orden=CATEGORIAS_ORDEN)


@app.route('/catalogo/<int:producto_id>/editar', methods=['GET', 'POST'])
def producto_editar(producto_id):
    """Editar un producto del catálogo. UPDATE conserva el id, así que NO toca
    los pedidos ni entregas guardados (que apuntan a este producto por su id)."""
    conn = get_db()
    producto = conn.execute("SELECT * FROM productos WHERE id = ?", (producto_id,)).fetchone()
    if not producto:
        conn.close()
        flash('No se encontró ese producto.', 'warning')
        return redirect(url_for('catalogo'))

    if request.method == 'POST':
        datos = _leer_form_producto()
        if not datos.get('descripcion') or not datos.get('codigo'):
            conn.close()
            flash('El código y la descripción son obligatorios.', 'warning')
            return render_template('producto_form.html', producto={**dict(producto), **datos},
                                   es_nuevo=False, categorias_display=CATEGORIAS_DISPLAY,
                                   categorias_orden=CATEGORIAS_ORDEN)
        set_clause = ', '.join(f"{c} = ?" for c in CAMPOS_PRODUCTO)
        valores = [datos[c] for c in CAMPOS_PRODUCTO] + [producto_id]
        conn.execute(f"UPDATE productos SET {set_clause} WHERE id = ?", valores)
        conn.commit()
        conn.close()
        flash(f"Producto «{datos['descripcion']}» actualizado.", 'success')
        return redirect(url_for('catalogo', q=datos.get('sku_tu_textil') or datos['codigo']))

    conn.close()
    return render_template('producto_form.html', producto=dict(producto), es_nuevo=False,
                           categorias_display=CATEGORIAS_DISPLAY,
                           categorias_orden=CATEGORIAS_ORDEN)


@app.route('/catalogo/<int:producto_id>/estado', methods=['POST'])
def producto_estado(producto_id):
    """Discontinuar o reactivar un producto (soft delete). No borra la fila,
    así los pedidos que lo referencian siguen funcionando."""
    nuevo = 'discontinuado' if request.form.get('accion') == 'discontinuar' else 'activo'
    conn = get_db()
    conn.execute("UPDATE productos SET estado = ? WHERE id = ?", (nuevo, producto_id))
    conn.commit()
    conn.close()
    if nuevo == 'discontinuado':
        flash('Producto discontinuado (oculto del catálogo, pero los pedidos siguen intactos).', 'info')
    else:
        flash('Producto reactivado.', 'success')
    return redirect(request.referrer or url_for('catalogo'))


# Campos que se pueden aplicar masivamente (los que tiene sentido compartir entre varios)
CAMPOS_MASIVOS = ['sku_tu_textil', 'mla_ids', 'categoria', 'linea', 'diseno', 'origen', 'coleccion']


@app.route('/catalogo/editar-masivo', methods=['POST'])
def producto_editar_masivo():
    """Aplica la misma info a varios productos a la vez. Solo cambia los campos
    que se marcaron con 'aplicar_<campo>'. Es UPDATE: no toca pedidos ni entregas."""
    ids = [int(i) for i in request.form.getlist('ids') if i.strip().isdigit()]
    if not ids:
        flash('No seleccionaste ningún producto.', 'warning')
        return redirect(request.referrer or url_for('catalogo'))

    sets, valores = [], []
    for campo in CAMPOS_MASIVOS:
        if request.form.get(f'aplicar_{campo}'):
            val = (request.form.get(campo) or '').strip()
            if campo == 'mla_ids' and val:
                partes = re.split(r'[\s,]+', val)
                val = ','.join(p for p in partes if p)
            sets.append(f"{campo} = ?")
            valores.append(val or None)

    estado_accion = request.form.get('estado_accion')
    if estado_accion == 'discontinuar':
        sets.append("estado = 'discontinuado'")
    elif estado_accion == 'reactivar':
        sets.append("estado = 'activo'")

    if not sets:
        flash('No marcaste ningún campo para cambiar.', 'warning')
        return redirect(request.referrer or url_for('catalogo'))

    placeholders = ','.join('?' for _ in ids)
    conn = get_db()
    conn.execute(
        f"UPDATE productos SET {', '.join(sets)} WHERE id IN ({placeholders})",
        valores + ids
    )
    conn.commit()
    conn.close()
    flash(f'{len(ids)} producto(s) actualizados a la vez.', 'success')
    return redirect(request.referrer or url_for('catalogo'))


# ── ENTREGAS (confirmaciones del proveedor) ──────────────────────────

def _skus_disponibles():
    """Lista de SKUs existentes con un nombre de modelo legible, para asignar a mano."""
    conn = get_db()
    rows = conn.execute("""
        SELECT sku_tu_textil, categoria, MIN(descripcion) as desc, COUNT(*) as variantes
        FROM productos
        WHERE sku_tu_textil IS NOT NULL
        GROUP BY sku_tu_textil
        ORDER BY categoria, sku_tu_textil
    """).fetchall()
    conn.close()
    skus = []
    for r in rows:
        # Nombre de modelo: descripción sin el color/código final
        nombre = r['desc'] or ''
        nombre = re.sub(r'\s+\d{4}\s*$', '', nombre).strip()  # quitar código de color
        nombre = re.sub(r'\s+\S+$', '', nombre).strip() if nombre else nombre  # quitar última palabra (color)
        cat = CATEGORIAS_DISPLAY.get(r['categoria'], r['categoria'] or '')
        skus.append({
            'sku': r['sku_tu_textil'],
            'nombre': nombre.title() if nombre else r['sku_tu_textil'],
            'categoria': cat,
            'label': f"{r['sku_tu_textil']} — {nombre.title()} ({cat})",
        })
    return skus


def _match_items_catalogo(items):
    """Para cada item parseado, busca el producto en el catálogo por código.
    Agrega producto_id, sku y nombre_catalogo (None si no matchea)."""
    conn = get_db()
    for it in items:
        prod = conn.execute(
            "SELECT id, descripcion, sku_tu_textil, categoria FROM productos WHERE codigo = ?",
            (it['codigo'],)
        ).fetchone()
        if prod:
            it['producto_id'] = prod['id']
            it['sku'] = prod['sku_tu_textil']
            it['categoria'] = prod['categoria']
            it['en_catalogo'] = True
        else:
            it['producto_id'] = None
            it['sku'] = None
            it['categoria'] = None
            it['en_catalogo'] = False
    conn.close()
    return items


@app.route('/entregas')
def entregas():
    """Lista de entregas/confirmaciones del proveedor, con filtros."""
    destino = request.args.get('destino', '').strip()
    periodo = request.args.get('periodo', '').strip()  # formato MM/AAAA sobre fecha_emision

    conn = get_db()
    conds = []
    params = []
    if destino:
        conds.append("e.destino = ?")
        params.append(destino)
    if periodo:
        # fecha_emision viene como dd/mm/aaaa → comparar el mm/aaaa final
        conds.append("substr(e.fecha_emision, 4) = ?")
        params.append(periodo)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    filas = conn.execute(f"""
        SELECT e.*,
            (SELECT COUNT(*) FROM entrega_items WHERE entrega_id = e.id) as n_items,
            (SELECT SUM(cantidad) FROM entrega_items WHERE entrega_id = e.id) as total_unidades
        FROM entregas e
        {where}
        ORDER BY e.fecha_emision DESC, e.created_at DESC
    """, params).fetchall()

    # Períodos disponibles para el filtro (mm/aaaa distintos)
    periodos = conn.execute("""
        SELECT DISTINCT substr(fecha_emision, 4) as periodo FROM entregas
        WHERE fecha_emision IS NOT NULL AND fecha_emision != ''
        ORDER BY substr(fecha_emision, 7) DESC, substr(fecha_emision, 4, 2) DESC
    """).fetchall()
    conn.close()
    return render_template('entregas.html', entregas=filas,
                           filtro_destino=destino, filtro_periodo=periodo,
                           periodos=[p['periodo'] for p in periodos])


@app.route('/entregas/nueva', methods=['GET', 'POST'])
def nueva_entrega():
    """Pegar el email del proveedor y previsualizar lo parseado."""
    if request.method == 'POST':
        texto = request.form.get('texto_email', '').strip()
        if not texto:
            flash('Pegá el texto del email de confirmación', 'warning')
            return render_template('nueva_entrega.html', preview=None)

        datos = parse_email(texto)
        datos['items'] = _match_items_catalogo(datos['items'])
        datos['texto_email'] = texto
        # Total de unidades
        datos['total_unidades'] = sum(it['cantidad'] or 0 for it in datos['items'])
        datos['n_en_catalogo'] = sum(1 for it in datos['items'] if it['en_catalogo'])

        return render_template('nueva_entrega.html',
                               preview=datos,
                               items_json=json.dumps(datos['items']),
                               skus_disponibles=_skus_disponibles(),
                               meses=MESES)

    return render_template('nueva_entrega.html', preview=None)


@app.route('/entregas/guardar', methods=['POST'])
def guardar_entrega():
    """Guardar la entrega parseada (viene del preview)."""
    items = json.loads(request.form.get('items_json', '[]'))
    destino = request.form.get('destino', '').strip() or None
    fecha_entrega = request.form.get('fecha_entrega', '').strip() or None
    notas = request.form.get('notas', '').strip()

    conn = get_db()
    cursor = conn.execute("""
        INSERT INTO entregas (destino, cliente, factura, remito, pedido_proveedor,
                              fecha_emision, fecha_vencimiento, fecha_entrega,
                              valor_total, texto_email, notas)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        destino,
        request.form.get('cliente', '').strip() or None,
        request.form.get('factura', '').strip() or None,
        request.form.get('remito', '').strip() or None,
        request.form.get('pedido_proveedor', '').strip() or None,
        request.form.get('fecha_emision', '').strip() or None,
        request.form.get('fecha_vencimiento', '').strip() or None,
        fecha_entrega,
        float(request.form['valor_total']) if request.form.get('valor_total') else None,
        request.form.get('texto_email', '').strip() or None,
        notas,
    ))
    entrega_id = cursor.lastrowid

    for idx, it in enumerate(items):
        # SKU asignado a mano para items que no matchearon el catálogo
        sku_manual = request.form.get(f'sku_item_{idx}', '').strip() or None
        conn.execute("""
            INSERT INTO entrega_items (entrega_id, producto_id, codigo_item, sku_manual, descripcion,
                                       cantidad, valor_unit, valor_total)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            entrega_id, it.get('producto_id'), it.get('codigo'), sku_manual, it.get('descripcion'),
            it.get('cantidad'), it.get('valor_unit'), it.get('valor_total'),
        ))

    conn.commit()
    conn.close()
    flash('Entrega guardada', 'success')
    return redirect(url_for('ver_entrega', entrega_id=entrega_id))


@app.route('/entregas/<int:entrega_id>')
def ver_entrega(entrega_id):
    """Detalle de una entrega."""
    conn = get_db()
    entrega = conn.execute("SELECT * FROM entregas WHERE id = ?", (entrega_id,)).fetchone()
    if not entrega:
        flash('Entrega no encontrada', 'error')
        return redirect(url_for('entregas'))

    items = conn.execute("""
        SELECT ei.*,
               COALESCE(p.sku_tu_textil, ei.sku_manual) as prod_sku,
               COALESCE(p.categoria, p3.categoria) as prod_categoria
        FROM entrega_items ei
        LEFT JOIN productos p ON ei.producto_id = p.id
        LEFT JOIN productos p3 ON p3.sku_tu_textil = ei.sku_manual
        WHERE ei.entrega_id = ?
        ORDER BY prod_categoria, ei.descripcion
    """, (entrega_id,)).fetchall()
    total = sum(i['cantidad'] or 0 for i in items)
    conn.close()
    return render_template('ver_entrega.html', entrega=entrega, items=items,
                           total_unidades=total, categorias_display=CATEGORIAS_DISPLAY)


@app.route('/entregas/<int:entrega_id>/exportar')
def exportar_entrega(entrega_id):
    """Exportar una entrega a Excel (cabecera + productos)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    conn = get_db()
    entrega = conn.execute("SELECT * FROM entregas WHERE id = ?", (entrega_id,)).fetchone()
    if not entrega:
        flash('Entrega no encontrada', 'error')
        return redirect(url_for('entregas'))
    items = conn.execute("""
        SELECT ei.*, p.sku_tu_textil as prod_sku
        FROM entrega_items ei
        LEFT JOIN productos p ON ei.producto_id = p.id
        WHERE ei.entrega_id = ?
        ORDER BY ei.descripcion
    """, (entrega_id,)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Entrega {entrega['factura'] or entrega_id}"[:31]
    thin = Border(*(Side(style='thin'),) * 4)

    dep = 'TUCUMÁN - SERGIO' if entrega['destino'] == 'tucuman' else 'ALSINA - ALBERTO'
    color = '548235' if entrega['destino'] == 'tucuman' else '2F5496'

    ws.merge_cells('A1:F1')
    ws['A1'] = f"ENTREGA COTEMINAS — {dep}"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')

    fent = fecha_legible_filter(entrega['fecha_entrega'])
    info = [
        ('Cliente', entrega['cliente'] or '-'),
        ('Factura', entrega['factura'] or '-'),
        ('Remito', entrega['remito'] or '-'),
        ('Pedido proveedor', entrega['pedido_proveedor'] or '-'),
        ('Emisión', entrega['fecha_emision'] or '-'),
        ('Fecha de entrega', fent or 'sin definir'),
    ]
    r = 2
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    r += 1
    headers = ['Código', 'SKU', 'Descripción', 'Cantidad', 'Valor unit.', 'Valor total']
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        c.border = thin
    r += 1
    total_u = 0
    for it in items:
        ws.cell(row=r, column=1, value=it['codigo_item'])
        ws.cell(row=r, column=2, value=it['prod_sku'] or '')
        ws.cell(row=r, column=3, value=it['descripcion'])
        ws.cell(row=r, column=4, value=it['cantidad'])
        ws.cell(row=r, column=5, value=it['valor_unit'])
        ws.cell(row=r, column=6, value=it['valor_total'])
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin
        total_u += it['cantidad'] or 0
        r += 1
    ws.cell(row=r, column=3, value='TOTAL UNIDADES:').font = Font(bold=True)
    ws.cell(row=r, column=4, value=total_u).font = Font(bold=True, size=12)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_dep = 'Tucuman' if entrega['destino'] == 'tucuman' else 'Alsina'
    filename = f"Entrega Coteminas - {nombre_dep} - Factura {entrega['factura'] or entrega_id}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/entregas/<int:entrega_id>/editar', methods=['POST'])
def editar_entrega(entrega_id):
    """Editar fecha de entrega, destino y notas de una entrega."""
    conn = get_db()
    conn.execute("""
        UPDATE entregas SET fecha_entrega = ?, destino = ?, notas = ?
        WHERE id = ?
    """, (
        request.form.get('fecha_entrega', '').strip() or None,
        request.form.get('destino', '').strip() or None,
        request.form.get('notas', '').strip(),
        entrega_id,
    ))
    conn.commit()
    conn.close()
    flash('Entrega actualizada', 'success')
    return redirect(url_for('ver_entrega', entrega_id=entrega_id))


@app.route('/entregas/<int:entrega_id>/eliminar', methods=['POST'])
def eliminar_entrega(entrega_id):
    """Eliminar una entrega."""
    conn = get_db()
    conn.execute("DELETE FROM entrega_items WHERE entrega_id = ?", (entrega_id,))
    conn.execute("DELETE FROM entregas WHERE id = ?", (entrega_id,))
    conn.commit()
    conn.close()
    flash('Entrega eliminada', 'info')
    return redirect(url_for('entregas'))


# ── COMPARATIVA: pedido vs confirmado ────────────────────────────────

@app.route('/comparativa')
def comparativa():
    """Cruza lo PEDIDO (de los pedidos) con lo CONFIRMADO (de las entregas),
    agrupado por SKU y depósito. Muestra cuánto pediste vs cuánto te llega y la diferencia."""
    destino_filtro = request.args.get('destino', '').strip()
    solo_finalizados = request.args.get('finalizados', '') == '1'

    conn = get_db()

    # ── PEDIDO: sumar pedido_items por SKU + destino ──
    cond_ped = []
    if destino_filtro:
        cond_ped.append("pi.destino = ?")
    if solo_finalizados:
        cond_ped.append("pe.estado = 'finalizado'")
    where_ped = ("WHERE " + " AND ".join(cond_ped)) if cond_ped else ""
    params_ped = [destino_filtro] if destino_filtro else []

    pedido_rows = conn.execute(f"""
        SELECT pi.destino,
               COALESCE(p.sku_tu_textil, pi.codigo_manual) as sku,
               COALESCE(p.categoria, pi.categoria_manual) as categoria,
               SUM(pi.cantidad) as cantidad
        FROM pedido_items pi
        JOIN pedidos pe ON pi.pedido_id = pe.id
        LEFT JOIN productos p ON pi.producto_id = p.id
        {where_ped}
        GROUP BY pi.destino, sku
    """, params_ped).fetchall()

    # ── CONFIRMADO: sumar entrega_items por SKU + destino ──
    cond_ent = []
    if destino_filtro:
        cond_ent.append("e.destino = ?")
    where_ent = ("WHERE " + " AND ".join(cond_ent)) if cond_ent else ""
    params_ent = [destino_filtro] if destino_filtro else []

    # Nota: NO unir por sku_manual (sku_tu_textil no es único → multiplicaría filas).
    # La categoría se resuelve después con sku_nombre.
    entrega_rows = conn.execute(f"""
        SELECT e.destino,
               COALESCE(p.sku_tu_textil, ei.sku_manual, p2.sku_tu_textil) as sku,
               COALESCE(p.categoria, p2.categoria) as categoria,
               SUM(ei.cantidad) as cantidad
        FROM entrega_items ei
        JOIN entregas e ON ei.entrega_id = e.id
        LEFT JOIN productos p ON ei.producto_id = p.id
        LEFT JOIN productos p2 ON p2.codigo = ei.codigo_item
        {where_ent}
        GROUP BY e.destino, sku
    """, params_ent).fetchall()

    # Nombre legible por SKU (descripción base de la familia)
    sku_nombre = {}
    for r in conn.execute("""
        SELECT sku_tu_textil, MIN(descripcion) as desc, categoria
        FROM productos WHERE sku_tu_textil IS NOT NULL GROUP BY sku_tu_textil
    """).fetchall():
        sku_nombre[r['sku_tu_textil']] = (r['desc'], r['categoria'])
    conn.close()

    # ── Combinar por (destino, sku) ──
    comb = {}  # (destino, sku) -> {pedido, confirmado, categoria}
    for r in pedido_rows:
        key = (r['destino'], r['sku'] or '(sin SKU)')
        comb.setdefault(key, {'pedido': 0, 'confirmado': 0, 'categoria': r['categoria']})
        comb[key]['pedido'] += r['cantidad'] or 0
    for r in entrega_rows:
        key = (r['destino'], r['sku'] or '(sin SKU)')
        comb.setdefault(key, {'pedido': 0, 'confirmado': 0, 'categoria': r['categoria']})
        comb[key]['confirmado'] += r['cantidad'] or 0
        if not comb[key]['categoria']:
            comb[key]['categoria'] = r['categoria']

    # Armar filas por destino
    resultado = {'tucuman': [], 'alsina': []}
    for (destino, sku), v in comb.items():
        if destino not in resultado:
            continue
        nombre, cat = sku_nombre.get(sku, (None, v['categoria']))
        # nombre legible: usar descripción del catálogo recortada, o el SKU
        if nombre:
            # sacar el color final para mostrar el modelo
            nombre_modelo = re.sub(r'\s+\d{4}\s*$', '', nombre).strip()
        else:
            nombre_modelo = sku
        resultado[destino].append({
            'sku': sku,
            'nombre': nombre_modelo,
            'categoria': cat or 'otro',
            'pedido': v['pedido'],
            'confirmado': v['confirmado'],
            'diferencia': v['pedido'] - v['confirmado'],
        })

    # Ordenar cada destino por categoría y nombre
    orden_cat = {c: i for i, c in enumerate(CATEGORIAS_ORDEN)}
    for d in resultado:
        resultado[d].sort(key=lambda x: (orden_cat.get(x['categoria'], 99), x['nombre'] or ''))

    return render_template('comparativa.html',
                           resultado=resultado,
                           filtro_destino=destino_filtro,
                           solo_finalizados=solo_finalizados,
                           categorias_display=CATEGORIAS_DISPLAY)


if __name__ == '__main__':
    init_db()
    print("\n" + "=" * 50)
    print("  PEDIDOS COTEMINAS - Tu Textil")
    print("  Abrir: http://localhost:5050")
    print("=" * 50 + "\n")
    app.run(debug=True, port=5050, use_reloader=False)
