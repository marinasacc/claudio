"""
Mapear SKU internos de Tu Textil y MLA de MercadoLibre a los productos del catálogo.

Lee el Excel PEDIDO COTEMINAS.xlsx y asigna los SKU y MLA a los productos
correspondientes en la base de datos.

Uso:
    python mapear_sku_mla.py
"""
import re
import openpyxl
from models import get_db

PEDIDO_PATH = r"G:\Mi unidad\ML - MARINA NO TOCAR\PEDIDOS PROVEEDORES\PEDIDO COTEMINAS.xlsx"


def extraer_items_pedido(filepath):
    """Extraer todos los items del Excel de pedido con sus MLA y SKU."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Hoja1']

    items = []
    skip_skus = {
        'SKU', 'TOALLAS', 'TOALLONES', 'JUEGOS', 'INSTITUCIONAL',
        'REPASADORES', 'TOALLA PISO', 'JUEGOS DE SABANAS',
        'PARA ARMAR JUEGOS (ESTO SE MANDA EN UN REMITO APARTE)', ''
    }

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        mla_raw = str(row[0].value).strip() if row[0].value else ''
        sku = str(row[1].value).strip() if row[1].value else ''
        desc = str(row[2].value).strip() if row[2].value else ''
        color = str(row[3].value).strip() if row[3].value else ''

        if not sku or sku in skip_skus:
            continue

        # Parsear MLAs (pueden ser múltiples separados por espacio)
        mla_ids = []
        if mla_raw:
            for part in mla_raw.split():
                part = part.strip()
                # MLA puede ser numérico puro o empezar con MLA
                if part.startswith('MLA'):
                    mla_ids.append(part)
                elif part.isdigit() and len(part) >= 6:
                    mla_ids.append(f"MLA{part}")

        items.append({
            'mla_ids': mla_ids,
            'sku': sku,
            'descripcion': desc,
            'color': color,
        })

    return items


def build_mapping_rules(items):
    """Construir reglas de mapeo: SKU -> (keywords para buscar en catálogo, MLA ids)."""
    # Agrupar por SKU para juntar todos los MLA
    sku_data = {}
    for item in items:
        sku = item['sku']
        if sku not in sku_data:
            sku_data[sku] = {
                'descripcion': item['descripcion'],
                'color': item['color'],
                'mla_ids': set(),
            }
        for mla in item['mla_ids']:
            sku_data[sku]['mla_ids'].add(mla)

    # Mapeo manual basado en el análisis de las descripciones del pedido
    # vs las descripciones del catálogo del proveedor
    # SKU -> (query SQL conditions, params)
    rules = {
        # TOALLAS
        'TOA171C': {
            'query': "categoria = 'toalla' AND linea = 'Fantasia' AND descripcion NOT LIKE '%RUFA%'",
            'desc': 'Toalla Fantasia 360gr 40x70'
        },
        'TOA141C': {
            'query': "categoria = 'toalla' AND linea = 'Prata' AND descripcion NOT LIKE '%CHEVRON%'",
            'desc': 'Toalla Prata 380gr 40x70'
        },
        'PA161C': {
            'query': "categoria = 'toalla' AND linea = 'Palette Urb'",
            'desc': 'Toalla Palette Urban 420gr 50x80'
        },
        'TOA140C': {
            'query': "categoria = 'toalla' AND descripcion LIKE '%BELLY%'",
            'desc': 'Toalla Arco Iris Belly 450gr 50x80'
        },
        'GR140C': {
            'query': "categoria = 'toalla' AND (descripcion LIKE '%DETROIT%' OR descripcion LIKE '%LIMA%' OR descripcion LIKE '%LUKKA%') AND fuente = 'arcoiris'",
            'desc': 'Toalla Arco Iris 500gr 50x80'
        },
        'PA140C': {
            'query': "categoria = 'toalla' AND linea = 'Palette Acc'",
            'desc': 'Toalla Palette Accent 500gr 50x80'
        },

        # TOALLONES
        'TOA171B': {
            'query': "categoria = 'toallon' AND linea = 'Fantasia'",
            'desc': 'Toallon Fantasia 360gr 70x130'
        },
        'TOA141B': {
            'query': "categoria = 'toallon' AND linea = 'Prata'",
            'desc': 'Toallon Prata 380gr 135x70'
        },
        'PA161B': {
            'query': "categoria = 'toallon' AND linea = 'Palette Urb'",
            'desc': 'Toallon Palette Urban 420gr'
        },
        'TOA140B': {
            'query': "categoria = 'toallon' AND descripcion LIKE '%BELLY%'",
            'desc': 'Toallon Arco Iris Belly 450gr'
        },
        'GR140B': {
            'query': "categoria = 'toallon' AND (descripcion LIKE '%DETROIT%' OR descripcion LIKE '%LIMA%' OR descripcion LIKE '%LUKKA%') AND fuente = 'arcoiris'",
            'desc': 'Toallon Arco Iris 500gr 150x80'
        },
        'GR140BG': {
            'query': "categoria = 'toallon_grande' AND fuente = 'arcoiris'",
            'desc': 'Toallon Grande Arco Iris 500gr 160x90'
        },
        'PA140B': {
            'query': "categoria = 'toallon' AND linea = 'Palette Acc'",
            'desc': 'Toallon Palette Accent 500gr'
        },

        # JUEGOS PARA ARMAR (toalla + toallon sueltos)
        'TOA171J': {
            'query': "categoria = 'juego' AND descripcion LIKE '%FANTASIA%'",
            'desc': 'Jgo Toalla y Toallon Fantasia'
        },
        'TOA141J': {
            'query': "categoria = 'juego' AND descripcion LIKE '%PRATA%'",
            'desc': 'Jgo Toalla y Toallon Prata'
        },
        'PA161J': {
            'query': "categoria = 'juego' AND descripcion LIKE '%PALETTE%' AND descripcion LIKE '%URB%'",
            'desc': 'Jgo Toalla y Toallon Palette 420'
        },
        'TOA140J': {
            'query': "categoria = 'juego' AND descripcion LIKE '%BELLY%'",
            'desc': 'Jgo Toalla y Toallon Arco Iris 450'
        },
        'GR140J': {
            'query': "categoria = 'juego' AND descripcion LIKE '%ARCO%' AND descripcion NOT LIKE '%BELLY%' AND descripcion NOT LIKE '%PALETTE%'",
            'desc': 'Jgo Toalla y Toallon Arco Iris 500'
        },

        # REPASADORES
        'GR115': {
            'query': "categoria = 'repasador' AND linea = 'Arcoiris'",
            'desc': 'Repasador Arco Iris'
        },
        'TOA114': {
            'query': "categoria = 'repasador' AND linea = 'Fantasia'",
            'desc': 'Repasador Fantasia'
        },
        'PA116': {
            'query': "categoria = 'repasador' AND linea = 'Palette Urb'",
            'desc': 'Repasador Palette'
        },

        # PISO
        'TOA151PIN': {
            'query': "categoria = 'piso' AND descripcion LIKE '%FANTASIA%'",
            'desc': 'Piso Fantasia Institucional'
        },
        'TOA140PIN': {
            'query': "categoria = 'piso' AND descripcion LIKE '%ARCO%'",
            'desc': 'Piso Arco Iris Institucional'
        },
        'PA140P': {
            'query': "categoria = 'piso' AND descripcion LIKE '%PALETTE%'",
            'desc': 'Piso Palette'
        },

        # SABANAS
        'GR181': {
            'query': "categoria = 'sabana' AND descripcion LIKE '%ARCO%' AND descripcion LIKE '%TWIN%'",
            'desc': 'Jgo Sabanas Arco Iris Twin'
        },
        'GR182': {
            'query': "categoria = 'sabana' AND descripcion LIKE '%ARCO%' AND (descripcion LIKE '%2 1/2%' OR descripcion LIKE '%FULL%')",
            'desc': 'Jgo Sabanas Arco Iris Full/2½'
        },
        'GR183': {
            'query': "categoria = 'sabana' AND descripcion LIKE '%ARCO%' AND descripcion LIKE '%QUEEN%'",
            'desc': 'Jgo Sabanas Arco Iris Queen'
        },
        'TOA3031': {
            'query': "categoria = 'sabana' AND linea = 'Prata' AND descripcion LIKE '%1 1/2%'",
            'desc': 'Jgo Sabanas Prata 1 Plaza'
        },
        'TOA3032': {
            'query': "categoria = 'sabana' AND linea = 'Prata' AND descripcion LIKE '%2 1/2%'",
            'desc': 'Jgo Sabanas Prata 2 Plazas'
        },
        'TOA3033': {
            'query': "categoria = 'sabana' AND linea = 'Prata' AND descripcion LIKE '%QUEEN%'",
            'desc': 'Jgo Sabanas Prata Queen'
        },
        'TOA3034': {
            'query': "categoria = 'sabana' AND linea = 'Prata' AND descripcion LIKE '%KING%'",
            'desc': 'Jgo Sabanas Prata King'
        },
    }

    # Combinar MLA ids con las rules
    for sku, data in rules.items():
        if sku in sku_data:
            data['mla_ids'] = sorted(sku_data[sku]['mla_ids'])
        else:
            data['mla_ids'] = []

    # Agregar SKUs que no tienen regla manual pero tienen MLA
    for sku, data in sku_data.items():
        if sku not in rules and data['mla_ids']:
            print(f"  AVISO: SKU '{sku}' tiene MLA {data['mla_ids']} pero no tiene regla de mapeo")
            print(f"         Desc: {data['descripcion']}")

    return rules, sku_data


def aplicar_mapeo(rules):
    """Aplicar el mapeo a la base de datos."""
    conn = get_db()

    # Agregar columnas si no existen
    try:
        conn.execute("ALTER TABLE productos ADD COLUMN sku_tu_textil TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE productos ADD COLUMN mla_ids TEXT")
    except Exception:
        pass

    # Limpiar mapeos anteriores
    conn.execute("UPDATE productos SET sku_tu_textil = NULL, mla_ids = NULL")

    total_mapeados = 0
    for sku, data in rules.items():
        query = data['query']
        mla_str = ','.join(data['mla_ids']) if data['mla_ids'] else None

        # Contar cuántos matchean
        count = conn.execute(f"SELECT COUNT(*) as cnt FROM productos WHERE {query}").fetchone()['cnt']

        if count > 0:
            conn.execute(f"""
                UPDATE productos SET sku_tu_textil = ?, mla_ids = ?
                WHERE {query}
            """, (sku, mla_str))
            total_mapeados += count
            mla_info = f" | MLA: {','.join(data['mla_ids'])}" if data['mla_ids'] else ""
            print(f"  OK  {sku:15s} -> {count:3d} productos | {data['desc']}{mla_info}")
        else:
            print(f"  --  {sku:15s} ->   0 productos | {data['desc']} (sin match en catalogo)")

    conn.commit()

    # También mapear items sin regla específica buscando en MLA que tengan datos
    # de juegos empaquetados, institucional, etc.

    # Resumen
    print(f"\nTotal productos mapeados: {total_mapeados}")

    # Stats
    con_sku = conn.execute("SELECT COUNT(*) as cnt FROM productos WHERE sku_tu_textil IS NOT NULL").fetchone()['cnt']
    con_mla = conn.execute("SELECT COUNT(*) as cnt FROM productos WHERE mla_ids IS NOT NULL AND mla_ids != ''").fetchone()['cnt']
    total = conn.execute("SELECT COUNT(*) as cnt FROM productos").fetchone()['cnt']
    print(f"Productos con SKU Tu Textil: {con_sku}/{total}")
    print(f"Productos con MLA: {con_mla}/{total}")

    # Listar SKUs únicos mapeados
    print("\nSKUs mapeados:")
    for row in conn.execute("""
        SELECT sku_tu_textil, mla_ids, COUNT(*) as cnt,
               GROUP_CONCAT(DISTINCT categoria) as cats
        FROM productos
        WHERE sku_tu_textil IS NOT NULL
        GROUP BY sku_tu_textil
        ORDER BY sku_tu_textil
    """).fetchall():
        print(f"  {row['sku_tu_textil']:15s} | {row['cnt']:3d} prods | {row['cats']:20s} | MLA: {row['mla_ids'] or '-'}")

    conn.close()


def main():
    print("=" * 70)
    print("  MAPEO SKU Tu Textil + MLA MercadoLibre -> Catalogo Coteminas")
    print("=" * 70)

    print("\n1. Extrayendo datos del Excel de pedidos...")
    items = extraer_items_pedido(PEDIDO_PATH)
    print(f"   {len(items)} lineas encontradas")

    # Listar todos los MLA encontrados
    all_mla = set()
    for item in items:
        for mla in item['mla_ids']:
            all_mla.add(mla)
    print(f"   {len(all_mla)} MLA unicos encontrados")

    print("\n2. Construyendo reglas de mapeo...")
    rules, sku_data = build_mapping_rules(items)
    print(f"   {len(rules)} reglas definidas")
    print(f"   {len(sku_data)} SKUs unicos en el pedido")

    print("\n3. Aplicando mapeo a la base de datos...")
    aplicar_mapeo(rules)

    print("\nMapeo completado!")


if __name__ == '__main__':
    main()
