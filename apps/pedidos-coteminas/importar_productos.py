"""
Importar productos desde las planillas Excel del proveedor a la base de datos.

Planillas soportadas:
- PLANILLA COTEMINAS - COLECCION OI26: Toallas, Sábanas, Almohadas (marca Coteminas/Palette/Fantasia/Prata)
- Planilla Colección AI OI26: Productos Arco Iris

Uso:
    python importar_productos.py
"""
import re
import openpyxl
from models import get_db, init_db

# ── Mapeo de categorías ──────────────────────────────────────────────

CATEGORIA_KEYWORDS = {
    'REPASADOR': 'repasador',
    'JUEGO': 'juego',
    'TOALLON G': 'toallon_grande',
    'TOALLON': 'toallon',
    'TOALLA PISO': 'piso',
    'TOALLITA': 'toallita',
    'PISO': 'piso',
    'TOALLA': 'toalla',
    'BATA': 'bata',
    'VISITA': 'visita',
    'SABANA': 'sabana',
    'ACOLCHADO': 'acolchado',
    'ALMOHADA': 'almohada',
    'FUNDA': 'funda',
    'PROTECTOR': 'protector',
    'MANTEL': 'mantel',
    'CAMINO': 'camino',
    'CORTINA': 'cortina',
    'CUBRE': 'cubre',
    'COLCHA': 'colcha',
    'ARCOIRIS 300H': 'sabana',   # Sábanas Arco Iris 300 hilos
}

# Líneas conocidas del proveedor
LINEAS = [
    'PALETTE ACC', 'PALETTE URB', 'PALETTE IVO', 'PALETTE',
    'ARCOIRIS ETE', 'ARCOIRIS', 'ARCO IRIS',
    'FANTASIA', 'PRATA', 'UNIQUE', 'CALFAT',
    'KAMACOLOR', 'MICROPERCAL', 'INSTITUCIONAL',
]

# Tamaños de sábanas/cama
TAMANIOS_SABANA = [
    'KING', 'QUEEN', '2 1/2', '1 1/2', '1 PLAZA', '2 PLAZAS',
    'TWIN', 'FULL', 'XL', 'MED', 'AJUS',
]

# Tamaños de toallas (se extraen del peso/medida en la descripción)
TAMANIO_TOALLA_MAP = {
    '40 X 70': '40x70',
    '40X70': '40x70',
    '50 X 80': '50x80',
    '50X80': '50x80',
    '70 X 130': '70x130',
    '70X130': '70x130',
    '135 X 70': '70x135',
    '80 X 150': '80x150',
    '90 X 160': '90x160',
    '100 X 150': '100x150',
    '30X30': '30x30',
    '48X70': '48x70',
    '48 X 70': '48x70',
}


def detectar_categoria(descripcion):
    """Detectar categoría del producto a partir de la descripción."""
    desc_upper = descripcion.upper()
    for keyword, cat in CATEGORIA_KEYWORDS.items():
        if desc_upper.startswith(keyword):
            return cat
    # Fallback: buscar en cualquier parte
    for keyword, cat in CATEGORIA_KEYWORDS.items():
        if keyword in desc_upper:
            return cat
    return 'otro'


def detectar_linea(descripcion):
    """Detectar la línea/marca del producto."""
    desc_upper = descripcion.upper()
    for linea in LINEAS:
        if linea in desc_upper:
            return linea.title()
    return None


def _limpiar_desc_para_diseno(descripcion, linea):
    """Quitar categoría, línea, peso, tamaño y CM de la descripción.
    Devuelve el texto que queda (en mayúsculas)."""
    desc_upper = descripcion.upper()

    # Quitar la parte de categoría del inicio (probar las keywords más largas primero)
    for keyword in sorted(CATEGORIA_KEYWORDS, key=len, reverse=True):
        if desc_upper.startswith(keyword):
            desc_upper = desc_upper[len(keyword):].strip()
            break

    # Quitar la línea (usar las LINEAS originales, más largas primero)
    for ln in sorted(LINEAS, key=len, reverse=True):
        if desc_upper.startswith(ln):
            desc_upper = desc_upper[len(ln):].strip()
            break
    # Si no estaba al inicio, quitarla donde aparezca
    if linea:
        desc_upper = desc_upper.replace(linea.upper(), '').strip()

    # Quitar peso (ej: "360 GR", "420 G.", "500 GR.")
    desc_upper = re.sub(r'\d+\s*GR?\.?\s*', ' ', desc_upper)
    # Quitar "CM"
    desc_upper = re.sub(r'\bCM\b', ' ', desc_upper)
    # Quitar tamaños de toallas (ej: 50 X 80)
    for tam_key in TAMANIO_TOALLA_MAP:
        desc_upper = desc_upper.replace(tam_key, ' ')
    # Quitar tamaños de sábanas
    for tam in TAMANIOS_SABANA:
        desc_upper = re.sub(r'\b' + re.escape(tam) + r'\b', ' ', desc_upper)
    # Normalizar espacios
    desc_upper = re.sub(r'\s+', ' ', desc_upper).strip()
    return desc_upper


def extraer_diseno_color(descripcion, linea, categoria):
    """Extraer diseño, nombre de color y código de color con lógica posicional.

    Patrón típico del catálogo:
        [CATEGORIA] [LINEA] [DISEÑO] [COLOR...] [CODIGO 4 dígitos]
    Ejemplos:
        TOALLON ARCOIRIS DETROIT ARENA 8805     -> Detroit / Arena / 8805
        TOALLON ARCOIRIS DETROIT GRIS PLOMO 9962 -> Detroit / Gris Plomo / 9962
        TOALLA PALETTE ACC SERENA BLANCO 0001    -> Serena / Blanco / 0001
        SABANA PRATA 2 1/2 ROAN                  -> Roan / None / None

    Devuelve (diseno, color_nombre, color_codigo).
    """
    desc_upper = descripcion.upper()

    # 1. Detectar código de color (4 dígitos al final)
    color_codigo = None
    m = re.search(r'\s(\d{4})\s*$', desc_upper)
    if m:
        color_codigo = m.group(1)

    # 2. Limpiar categoría/línea/peso/tamaño/CM
    resto = _limpiar_desc_para_diseno(descripcion, linea)

    # 3. Quitar el código de color del resto si está
    if color_codigo:
        resto = re.sub(r'\s*' + color_codigo + r'\s*$', '', resto).strip()

    # 4. Quitar palabras genéricas
    tokens = [w for w in resto.split()
              if w not in ('DE', 'Y', 'EN', 'CON', 'PACK', 'PVC', 'X', 'INST', '-')]
    tokens = [w for w in tokens if not re.match(r'^\d+$', w)]

    if not tokens:
        return None, None, color_codigo

    if color_codigo:
        # Hay color con código: 1ra palabra = diseño, resto = nombre del color
        diseno = tokens[0]
        color_nombre = ' '.join(tokens[1:]) if len(tokens) > 1 else None
        return diseno.title(), (color_nombre.title() if color_nombre else None), color_codigo
    else:
        # Sin código de color: todo lo que queda es el diseño/modelo
        # (sábanas por modelo, almohadas, etc.)
        return ' '.join(tokens).title(), None, None


def extraer_tamanio(descripcion):
    """Extraer tamaño del producto."""
    desc_upper = descripcion.upper()

    # Tamaños de sábana
    for tam in TAMANIOS_SABANA:
        if tam in desc_upper:
            return tam.title()

    # Tamaños de toalla
    for tam_key, tam_val in TAMANIO_TOALLA_MAP.items():
        if tam_key in desc_upper:
            return tam_val

    return None


def importar_planilla_coteminas(filepath, coleccion='OI26'):
    """Importar productos de la planilla Coteminas (hoja TOTAL)."""
    print(f"\n📦 Importando planilla Coteminas: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb['TOTAL']

    productos = []
    seccion_actual = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        col_a = row[0].value  # Estado (nuevo/permanece) o separador
        col_b = row[1].value  # Item/Código
        col_c = row[2].value  # Descripción
        col_d = row[3].value  # Familia Comercial
        col_e = row[4].value  # Familia Material
        col_f = row[5].value  # Peso neto
        col_g = row[6].value  # Peso bruto
        col_h = row[7].value  # EAN

        # Detectar separadores de sección
        if col_a and isinstance(col_a, str) and not col_b:
            seccion_actual = col_a.strip()
            continue

        # Saltar filas sin código o sin descripción
        if not col_b or not col_c:
            continue
        if not isinstance(col_c, str):
            continue

        descripcion = col_c.strip()
        codigo = str(col_b).strip()

        # EAN como string sin .0
        ean = ''
        if col_h:
            ean = str(col_h).replace('.0', '').strip()

        peso = None
        if col_f:
            try:
                peso = float(col_f)
            except (ValueError, TypeError):
                pass

        categoria = detectar_categoria(descripcion)
        linea = detectar_linea(descripcion)
        diseno, color_nombre, color_codigo = extraer_diseno_color(descripcion, linea, categoria)
        tamanio = extraer_tamanio(descripcion)

        # Tipo (Jacquard, Lisa, etc.)
        tipo = None
        col_tipo_idx = 19  # columna T en hoja TOTAL
        if len(row) > col_tipo_idx and row[col_tipo_idx].value:
            tipo = str(row[col_tipo_idx].value).strip()

        # Origen
        origen = None
        col_origen_idx = 17  # columna R
        if len(row) > col_origen_idx and row[col_origen_idx].value:
            origen = str(row[col_origen_idx].value).strip()

        familia_comercial = str(col_d).strip() if col_d else None
        familia_material = str(col_e).strip() if col_e else None

        productos.append({
            'codigo': codigo,
            'descripcion': descripcion,
            'categoria': categoria,
            'linea': linea,
            'diseno': diseno,
            'color_nombre': color_nombre,
            'color_codigo': color_codigo,
            'tipo': tipo,
            'tamanio': tamanio,
            'ean': ean,
            'peso_neto': peso,
            'familia_comercial': familia_comercial,
            'familia_material': familia_material,
            'origen': origen,
            'coleccion': coleccion,
            'fuente': 'coteminas',
        })

    print(f"  Encontrados {len(productos)} productos")
    return productos


def importar_planilla_arcoiris(filepath, coleccion='OI26'):
    """Importar productos de la planilla Arco Iris."""
    print(f"\n🌈 Importando planilla Arco Iris: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    productos = []

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        # Estructura: A=num, B=Línea, C=estado, D=Codigo, E=Descripción, F=Fam.Mat, G=Fam.Com, H=Origen, I=Peso, J=Bruto, K=EAN
        col_linea = row[1].value
        col_estado = row[2].value
        col_codigo = row[3].value
        col_desc = row[4].value
        col_fam_mat = row[5].value
        col_fam_com = row[6].value
        col_origen = row[7].value
        col_peso = row[8].value
        col_ean = row[10].value if len(row) > 10 else None

        if not col_codigo or not col_desc:
            continue
        if not isinstance(col_desc, str):
            continue

        descripcion = col_desc.strip()
        codigo = str(col_codigo).strip()

        ean = ''
        if col_ean:
            ean = str(col_ean).replace('.0', '').strip()

        peso = None
        if col_peso:
            try:
                peso = float(col_peso)
            except (ValueError, TypeError):
                pass

        categoria = detectar_categoria(descripcion)
        linea = detectar_linea(descripcion)
        if not linea and col_linea:
            linea = str(col_linea).strip().title()
        diseno, color_nombre, color_codigo = extraer_diseno_color(descripcion, linea, categoria)
        tamanio = extraer_tamanio(descripcion)

        familia_comercial = str(col_fam_com).strip() if col_fam_com else None
        familia_material = str(col_fam_mat).strip() if col_fam_mat else None
        origen = str(col_origen).strip() if col_origen else None

        productos.append({
            'codigo': codigo,
            'descripcion': descripcion,
            'categoria': categoria,
            'linea': linea,
            'diseno': diseno,
            'color_nombre': color_nombre,
            'color_codigo': color_codigo,
            'tipo': None,
            'tamanio': tamanio,
            'ean': ean,
            'peso_neto': peso,
            'familia_comercial': familia_comercial,
            'familia_material': familia_material,
            'origen': origen,
            'coleccion': coleccion,
            'fuente': 'arcoiris',
        })

    print(f"  Encontrados {len(productos)} productos")
    return productos


def guardar_productos(productos):
    """Guardar productos en la base de datos. Limpia e inserta todo."""
    conn = get_db()

    # Contar existentes
    count = conn.execute("SELECT COUNT(*) FROM productos").fetchone()[0]
    if count > 0:
        print(f"\n⚠️  Ya hay {count} productos en la DB. Limpiando...")
        conn.execute("DELETE FROM productos")

    for p in productos:
        conn.execute("""
            INSERT INTO productos (
                codigo, descripcion, categoria, linea, diseno,
                color_nombre, color_codigo, tipo, tamanio, ean,
                peso_neto, familia_comercial, familia_material, origen,
                coleccion, fuente
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            p['codigo'], p['descripcion'], p['categoria'], p['linea'], p['diseno'],
            p['color_nombre'], p['color_codigo'], p['tipo'], p['tamanio'], p['ean'],
            p['peso_neto'], p['familia_comercial'], p['familia_material'], p['origen'],
            p['coleccion'], p['fuente'],
        ))

    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM productos").fetchone()[0]
    print(f"\n✅ Total importados: {total} productos")

    # Resumen por categoría
    print("\n📊 Resumen por categoría:")
    for row in conn.execute(
        "SELECT categoria, COUNT(*) as cnt FROM productos GROUP BY categoria ORDER BY cnt DESC"
    ):
        print(f"  {row['categoria']:20s} {row['cnt']:4d}")

    # Resumen por fuente
    print("\n📊 Resumen por fuente:")
    for row in conn.execute(
        "SELECT fuente, COUNT(*) as cnt FROM productos GROUP BY fuente ORDER BY cnt DESC"
    ):
        print(f"  {row['fuente']:20s} {row['cnt']:4d}")

    conn.close()


def main():
    # Rutas de las planillas
    PLANILLA_COTEMINAS = r"G:\Mi unidad\ML - MARINA NO TOCAR\CODIGOS DE BARRA\PLANILLA COTEMINAS - COLECCION OI26 (3).xlsx"
    PLANILLA_ARCOIRIS = r"G:\Mi unidad\ML - MARINA NO TOCAR\CODIGOS DE BARRA\Planilla Colección AI OI26 (2).xlsx"

    print("=" * 60)
    print("  IMPORTADOR DE PRODUCTOS COTEMINAS")
    print("=" * 60)

    # Inicializar DB
    init_db()

    # Importar ambas planillas
    productos_coteminas = importar_planilla_coteminas(PLANILLA_COTEMINAS)
    productos_arcoiris = importar_planilla_arcoiris(PLANILLA_ARCOIRIS)

    # Combinar y guardar
    todos = productos_coteminas + productos_arcoiris
    guardar_productos(todos)

    print("\n🎉 Importación completada!")


if __name__ == '__main__':
    main()
